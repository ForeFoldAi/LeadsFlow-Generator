"""
Lead engine service — background scraping + DB persistence + export helpers.

This module bridges the existing root-level orchestrator (unchanged) with
the new FastAPI / SQLAlchemy persistence layer.
"""
import io
import uuid
from datetime import datetime
from typing import List, Optional


# ── Background job ────────────────────────────────────────────────────────────

async def run_scrape_job(
    job_id: str,
    keyword: str,
    location: str,
    sources: List[str],
    max_per_source: int = 25,
    min_score: float = 0.0,
    country: str = "India",
    headless: bool = True,
    delay: float = 1.5,
) -> None:
    """
    Background coroutine: run all scrapers, score leads, persist to DB.
    Called by FastAPI BackgroundTasks — runs after the HTTP response is sent.
    """
    import re

    from app.core.database import AsyncSessionLocal
    from app.models.job import ScrapeJob
    from app.models.lead import Lead as LeadModel
    from sqlalchemy import select

    def _norm_phone(p: str) -> str:
        return re.sub(r"\D", "", p or "")

    def _norm_key(name: str, city: str) -> str:
        return re.sub(r"[^a-z0-9]", "", (name or "").lower()) + "|" + re.sub(r"[^a-z0-9]", "", (city or "").lower())

    async with AsyncSessionLocal() as db:
        # ── Mark job as running ───────────────────────────────────────────────
        res = await db.execute(select(ScrapeJob).where(ScrapeJob.id == job_id))
        job = res.scalar_one_or_none()
        if not job:
            return

        job.status = "running"
        job.updated_at = datetime.utcnow()
        await db.commit()

        try:
            from app.services.orchestrator import LeadEngine

            engine = LeadEngine(
                sources=sources,
                max_per_source=max_per_source,
                delay=delay,
                headless=headless,
                min_score=min_score,
                country=country,
            )
            leads = await engine.run(keyword, location)

            # ── Load existing leads for this user to detect duplicates ────────
            existing_phones: set = set()
            existing_name_city: set = set()
            if job.user_id:
                # Check 1: Generator's own scraped_leads table
                existing_q = (
                    select(LeadModel.phone, LeadModel.name, LeadModel.city)
                    .join(ScrapeJob, LeadModel.job_id == ScrapeJob.id)
                    .where(ScrapeJob.user_id == job.user_id)
                )
                existing_res = await db.execute(existing_q)
                for row in existing_res.all():
                    ph = _norm_phone(row.phone)
                    if ph:
                        existing_phones.add(ph)
                    key = _norm_key(row.name, row.city)
                    if key != "|":
                        existing_name_city.add(key)

                # Check 2: CRM leads table (manually added / imported leads)
                from sqlalchemy import text
                crm_q = text(
                    "SELECT phone_number, name, city FROM leads WHERE user_id = :uid"
                )
                crm_res = await db.execute(crm_q, {"uid": job.user_id})
                for row in crm_res.all():
                    ph = _norm_phone(row.phone_number)
                    if ph:
                        existing_phones.add(ph)
                    key = _norm_key(row.name, row.city)
                    if key != "|":
                        existing_name_city.add(key)

            # ── Persist scored leads (skip duplicates) ────────────────────────
            duplicate_count = 0
            for lead in leads:
                ph = _norm_phone(lead.phone)
                key = _norm_key(lead.name, lead.city)
                if job.user_id and (
                    (ph and ph in existing_phones)
                    or (key != "|" and key in existing_name_city)
                ):
                    duplicate_count += 1
                    continue

                db_lead = LeadModel(
                    id=str(uuid.uuid4()),
                    job_id=job_id,
                    name=lead.name,
                    category=lead.category,
                    phone=lead.phone,
                    website=lead.website,
                    email=lead.email,
                    address=lead.address,
                    city=lead.city,
                    state=lead.state,
                    zip_code=lead.zip_code,
                    hours=getattr(lead, 'hours', ''),
                    sources=lead.sources,
                    source_urls=lead.source_urls,
                    enriched_from=lead.enriched_from,
                    confidence_score=lead.confidence_score,
                    scraped_at=lead.scraped_at,
                    created_at=datetime.utcnow(),
                )
                db.add(db_lead)
                # Track newly added leads in-memory to prevent duplicates
                # within this same batch
                if ph:
                    existing_phones.add(ph)
                if key != "|":
                    existing_name_city.add(key)

            job.status = "completed"
            job.total_found = len(leads) - duplicate_count
            job.duplicate_count = duplicate_count
            job.completed_at = datetime.utcnow()
            job.updated_at = datetime.utcnow()
            await db.commit()

        except Exception as exc:
            job.status = "failed"
            job.error_message = str(exc)
            job.updated_at = datetime.utcnow()
            await db.commit()
            raise


# ── Export helpers ────────────────────────────────────────────────────────────

def leads_to_leadsflow_rows(
    db_leads,
    keyword: str = "",
    country: str = "India",
) -> List[dict]:
    """
    Convert a list of SQLAlchemy Lead models to LeadsFlow-format row dicts.
    Uses the existing orchestrator conversion logic verbatim.
    """
    from app.models.lead_dataclass import Lead as LeadDataclass
    from app.services.orchestrator import lead_to_leadsflow

    rows = []
    for dl in db_leads:
        dataclass_lead = LeadDataclass(
            name=dl.name or "",
            category=dl.category or "",
            phone=dl.phone or "",
            website=dl.website or "",
            email=dl.email or "",
            address=dl.address or "",
            city=dl.city or "",
            state=dl.state or "",
            zip_code=dl.zip_code or "",
            sources=dl.sources or "",
            source_urls=dl.source_urls or "",
            enriched_from=dl.enriched_from or "",
            confidence_score=dl.confidence_score or 0.0,
            scraped_at=dl.scraped_at or "",
        )
        row = lead_to_leadsflow(dataclass_lead, keyword=keyword, country=country)
        if row:
            rows.append(row)
    return rows


def build_excel_bytes(rows: List[dict]) -> bytes:
    """
    Build a color-coded Excel workbook from LeadsFlow rows and return raw bytes.
    """
    if not rows:
        return b""

    import pandas as pd
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from app.services.orchestrator import LEADSFLOW_COLUMNS

    df = pd.DataFrame(rows, columns=LEADSFLOW_COLUMNS)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="LeadsFlow Import")
        ws = writer.sheets["LeadsFlow Import"]

        # Header styling
        hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 42

        tier_fills = {
            "Hot":      PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "Strong":   PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "Good":     PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "Moderate": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            "Weak":     PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        }
        alt = [
            PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"),
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        ]
        bfont = Font(name="Arial", size=9)
        border = Border(bottom=Side(style="thin", color="DDDDDD"))
        notes_idx = LEADSFLOW_COLUMNS.index("Additional Notes")

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            fill = alt[i % 2]
            for cell in row:
                cell.fill = fill
                cell.font = bfont
                cell.border = border
                cell.alignment = Alignment(vertical="top")

        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(w + 3, 40)
        ws.freeze_panes = "A2"

    output.seek(0)
    return output.read()
